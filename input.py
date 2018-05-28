from openpyxl import load_workbook
import os

# GLOBAL VARIABLES
income_types = ['Loose Plate', 'Diezmo', 'Pro-Templo', 'Other']
file = "data.xlsx"
sheet = "2017_Income_Data"


# INSERTS INTO EXCEL SHEET BY INCOME TYPE
def insert_to_sheet(month, ins):

    # Setup, open sheet
    wb = load_workbook(file)
    ws = wb[sheet]

    insert_row = ws.max_row + 1         # Get row to insert new entry 
    
    # item is tuple of a type and amount (special case is Other that has note)
    for item in ins:
        type_in = item[0]               # Set type of income
        amount = item[1]                # Amount of the entry 
        other = False                   # Bool to detect Other field 
        if type_in ==  "Other":         # Wether 'note' field is required
            pod = "Other"
            note = item[2]
            other = True
        else:
            pod = "Diezmo y Ofrenda"

        # create object to insert easily
        temp_lst = [month, pod, type_in, amount]

        # Add note field 
        if other:
            temp_lst.append(note)

        # Insert 'to_col' number of cells in 'insert_row'
        for col in range(1,len(temp_lst)+1):
            ws.cell(column = col, row=insert_row, value=temp_lst[col-1])
        print("Inserted: " + str(temp_lst))

        insert_row = insert_row + 1


    # Save inserts
    wb.save("data.xlsx")


# GET REQUIRED FIELDS FOR ANY TYPE
def get_amount(typ):
    # Enter amount for specific income type
    print "Enter " + income_types[int(typ)-1] + " Amount:"
    amount = float(raw_input(">"))
    
    # Add note to describe an 'Other' amount
    if typ == str(len(income_types)):
        print ("Enter Note for Other:")
        note = raw_input(">")
        tup = (income_types[int(typ)-1], amount, note)
    else:
        tup = (income_types[int(typ)-1], amount)
    os.system('cls')
    return tup


# GET ALL INCOME FOR EITHER SERVICE OR PER MONTH 
def get_service_income(month):
    income = []

    # Display avaliable options
    while(True):
        print("Working in month: " + month)
        for item in income_types:
            print str(income_types.index(item) + 1) + ". " + item 
        print(str(len(income_types) + 1) + ". Exit" )
        type_of_input = raw_input("> ")
        os.system('cls')

        # select to continue or exit
        if type_of_input == str(len(income_types) + 1): # Last option will always be Exit
            break
        else:
            tup = get_amount(type_of_input)
            income.append(tup)
    
    # Enter into sheet 
    insert_to_sheet(month, income)


def get_entry():
    month = raw_input("Enter Month / Mes for Entry:\n>")
    os.system('cls')
    print "Current Month Chosen: " + month
    get_service_income(month)

    while(True):
        print("1. Change Month")
        print("2. Exit")
        choice = int(raw_input(">"))

        if (choice == 1):
            month = raw_input("Enter Month / Mes for Entry:\n>")
            os.system('cls')
            print "Current Month Chosen: " + month
            get_service_income(month)
        elif(choice == 2):
            break

    



if __name__ == "__main__":
    #print("")
    get_entry()
    #lst = [('Loose PLate', 32),('Diezmo',789),('Pro-Templo', 40), ('Other', 50, 'kids room')]
    #insert_to_sheet("Jan / Enero",lst)